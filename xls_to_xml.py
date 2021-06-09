import xml.etree.ElementTree as et
import openpyxl as xl
from openpyxl.utils import get_column_letter

flg = 0



wb = xl.load_workbook(filename='Sample2.xlsx')
ws = wb.active

print('Enter Company name')
comp_name = input()

headers = {}
for idx, cell in enumerate(ws.iter_cols(1, ws.max_column), start=0):
    headers[cell[0].value] = idx


root = et.Element('ENVELOPE')

HEADER = et.Element('HEADER')
root.append(HEADER)

TALLYREQUEST = et.SubElement(HEADER, 'TALLYREQUEST')
TALLYREQUEST.text = 'Import Data'

BODY = et.Element('BODY')
root.append(BODY)

IMPORTDATA = et.Element('IMPORTDATA')
BODY.append(IMPORTDATA)

REQUESTDESC = et.Element('REQUESTDESC')
IMPORTDATA.append(REQUESTDESC)

REPORTNAME = et.SubElement(REQUESTDESC, 'REPORTNAME')
REPORTNAME.text = 'Vouchers'

STATICVARIABLES = et.SubElement(REQUESTDESC, 'STATICVARIABLES')
SVCURRENTCOMPANY = et.SubElement(STATICVARIABLES, 'SVCURRENTCOMPANY')
SVCURRENTCOMPANY.text = comp_name

REQUESTDATA = et.Element('REQUESTDATA')
IMPORTDATA.append(REQUESTDATA)

for idx, row in enumerate(ws.values):

    def acttest(string):
        #print('In acttest header test:',headers.get(str(string)))
        if headers.get(str(string)) != None:
            if row[headers[str(string)]] != None and row[headers[str(string)]] != 0 :
                   return True
            else:
                return False
        else:
            return False
                
    #print(row)
    if idx != 0:
        TALLYMESSAGE = et.SubElement(REQUESTDATA, 'TALLYMESSAGE')
        TALLYMESSAGE.set('xmlns:UDF', 'TallyUDF')
        VOUCHER = et.SubElement(TALLYMESSAGE, 'VOUCHER', REMOTEID="", VCHKEY="", VCHTYPE="Purchase", ACTION="Create", OBJVIEW="Accounting Voucher View")
        b = et.SubElement(VOUCHER, 'OLDAUDITENTRYIDS.LIST', TYPE='Number')
        OLDAUDITENTRYIDS = et.SubElement(b, 'OLDAUDITENTRYIDS')
        OLDAUDITENTRYIDS.text = '-1'
        a = et.SubElement(VOUCHER, 'ADDRESS.LIST', TYPE='String')
        for m in range(1, 5):
            s = str('Address ' + str(m))
            if  acttest(s) == True:
                ADDRESS = et.SubElement(a, 'ADDRESS')
                ADDRESS.text = row[headers[s]]
                flg += 1
        if flg < 1:
            VOUCHER.remove(a)
        DATE = et.SubElement(VOUCHER, 'DATE')
        DATE.text = str(row[headers['Date']].strftime('%Y%m%d')) #date format here
        if acttest('Supplier Inv Date') == True:
            REFERENCEDATE = et.SubElement(VOUCHER, 'REFERENCEDATE')
            REFERENCEDATE.text = row[headers['Supplier Inv Date']].strftime('%y%m%d')
        if acttest('State') == True:
            STATENAME = et.SubElement(VOUCHER, 'STATENAME')
            STATENAME.text = row[headers['State']]
        if acttest('GST Registration Type') == True:       
            GSTREGISTRATIONTYPE = et.SubElement(VOUCHER, 'GSTREGISTRATIONTYPE')
            GSTREGISTRATIONTYPE.text = row[headers['GST Registration Type']]
        if acttest('Narration') == True:
            
            NARRATION = et.SubElement(VOUCHER, 'NARRATION')
            NARRATION.text = row[headers['Narration']]
        COUNTRYOFRESIDENCE = et.SubElement(VOUCHER, 'COUNTRYOFRESIDENCE')
        COUNTRYOFRESIDENCE.text = 'India'
        if acttest('GSTIN') == True:      
            PARTYGSTIN = et.SubElement(VOUCHER, 'PARTYGSTIN')
            PARTYGSTIN.text = row[headers['GSTIN']]
        PARTYNAME = et.SubElement(VOUCHER, 'PARTYNAME')
        PARTYNAME.text = row[5]
        VOUCHERTYPENAME = et.SubElement(VOUCHER, 'VOUCHERTYPENAME')
        VOUCHERTYPENAME.text = row[headers['Voucher Type']]
        REFERENCE = et.SubElement(VOUCHER, 'REFERENCE')
        REFERENCE.text = row[headers['Supplier Inv No']]
        VOUCHERNUMBER = et.SubElement(VOUCHER, 'VOUCHERNUMBER')
        VOUCHERNUMBER.text = row[headers['Voucher No']]
        PARTYLEDGERNAME = et.SubElement(VOUCHER, 'PARTYLEDGERNAME')
        PARTYLEDGERNAME.text = row[5]
        BASICBASEPARTYNAME = et.SubElement(VOUCHER, 'BASICBASEPARTYNAME')
        BASICBASEPARTYNAME.text = row[5]
        if acttest('CST No') == True:       
            BUYERSCSTNUMBER = et.SubElement(VOUCHER, 'BUYERSCSTNUMBER')
            BUYERSCSTNUMBER.text = row[headers['CST No']]
        PERSISTEDVIEW = et.SubElement(VOUCHER, 'PERSISTEDVIEW')
        PERSISTEDVIEW.text = 'Accounting Voucher View'
        BASICBUYERNAME = et.SubElement(VOUCHER, 'BASICBUYERNAME')
        BASICBUYERNAME.text = comp_name
        if acttest('VAT Tin No') == True:
            BASICBUYERSSALESTAXNO = et.SubElement(VOUCHER, 'BASICBUYERSSALESTAXNO')
            BASICBUYERSSALESTAXNO.text = row[headers['VAT Tin No']]
        BASICDATETIMEOFINVOICE = et.SubElement(VOUCHER, 'BASICDATETIMEOFINVOICE')
        BASICDATETIMEOFINVOICE.text = str(row[0].strftime('%d-%B-%y')) # enter date format heree
        BASICDATETIMEOFREMOVAL = et.SubElement(VOUCHER, 'BASICDATETIMEOFREMOVAL')
        BASICDATETIMEOFREMOVAL.text = str(row[0].strftime('%d-%B-%y')) # enter date fromat here
        if acttest('Cost Center') == True:
            COSTCENTRENAME = et.SubElement(VOUCHER, 'COSTCENTRENAME')
            COSTCENTRENAME.text = row[headers['Cost Center']]
        DIFFACTUALQTY = et.SubElement(VOUCHER, 'DIFFACTUALQTY')
        DIFFACTUALQTY.text = 'No'
        ISMSTFROMSYNC = et.SubElement(VOUCHER, 'ISMSTFROMSYNC')
        ISMSTFROMSYNC.text = 'No'
        ASORIGINAL = et.SubElement(VOUCHER, 'ASORIGINAL')
        ASORIGINAL.text = 'No'
        AUDITED = et.SubElement(VOUCHER, 'AUDITED')
        AUDITED.text = 'No'
        FORJOBCOSTING = et.SubElement(VOUCHER, 'FORJOBCOSTING')
        FORJOBCOSTING.text = 'No'
        ISOPTIONAL = et.SubElement(VOUCHER, 'ISOPTIONAL')
        ISOPTIONAL.text = 'No'
        if acttest('Supplier Inv Date') == True:
            EFFECTIVEDATE = et.SubElement(VOUCHER, 'EFFECTIVEDATE')
            EFFECTIVEDATE.text = row[headers['Supplier Inv Date']].strftime('%Y%m%d') # enter date format here
        USEFOREXCISE = et.SubElement(VOUCHER, 'USEFOREXCISE')
        USEFOREXCISE.text = 'No'
        ISFORJOBWORKIN = et.SubElement(VOUCHER, 'ISFORJOBWORKIN')
        ISFORJOBWORKIN.text = 'No'
        ALLOWCONSUMPTION = et.SubElement(VOUCHER, 'ALLOWCONSUMPTION')
        ALLOWCONSUMPTION.text = 'No'
        USEFORINTEREST = et.SubElement(VOUCHER, 'USEFORINTEREST')
        USEFORINTEREST.text = 'No'
        USEFORGAINLOSS = et.SubElement(VOUCHER, 'USEFORGAINLOSS')
        USEFORGAINLOSS.text = 'No'
        USEFORGODOWNTRANSFER = et.SubElement(VOUCHER, 'USEFORGODOWNTRANSFER')
        USEFORGODOWNTRANSFER.text = 'No'
        USEFORCOMPOUND = et.SubElement(VOUCHER, 'USEFORCOMPOUND')
        USEFORCOMPOUND.text = 'No'
        USEFORSERVICETAX = et.SubElement(VOUCHER, 'USEFORSERVICETAX')
        USEFORSERVICETAX.text = 'No'
        ISEXCISEVOUCHER = et.SubElement(VOUCHER, 'ISEXCISEVOUCHER')
        ISEXCISEVOUCHER.text = 'No'
        EXCISETAXOVERRIDE = et.SubElement(VOUCHER, 'EXCISETAXOVERRIDE')
        EXCISETAXOVERRIDE.text = 'No'
        USEFORTAXUNITTRANSFER = et.SubElement(VOUCHER, 'USEFORTAXUNITTRANSFER')
        USEFORTAXUNITTRANSFER.text = 'No'
        EXCISEOPENING = et.SubElement(VOUCHER, 'EXCISEOPENING')
        EXCISEOPENING.text = 'No'
        USEFORFINALPRODUCTION = et.SubElement(VOUCHER, 'USEFORFINALPRODUCTION')
        USEFORFINALPRODUCTION.text = 'No'
        ISTDSOVERRIDDEN = et.SubElement(VOUCHER, 'ISTDSOVERRIDDEN')
        ISTDSOVERRIDDEN.text = 'No'
        ISTCSOVERRIDDEN = et.SubElement(VOUCHER, 'ISTCSOVERRIDDEN')
        ISTCSOVERRIDDEN.text = 'No'
        ISTDSTCSCASHVCH = et.SubElement(VOUCHER, 'ISTDSTCSCASHVCH')
        ISTDSTCSCASHVCH.text = 'No'
        INCLUDEADVPYMTVCH = et.SubElement(VOUCHER, 'INCLUDEADVPYMTVCH')
        INCLUDEADVPYMTVCH.text = 'No'
        ISSUBWORKSCONTRACT = et.SubElement(VOUCHER, 'ISSUBWORKSCONTRACT')
        ISSUBWORKSCONTRACT.text = 'No'
        ISVATOVERRIDDEN = et.SubElement(VOUCHER, 'ISVATOVERRIDDEN')
        ISVATOVERRIDDEN.text = 'No'
        IGNOREORIGVCHDATE = et.SubElement(VOUCHER, 'IGNOREORIGVCHDATE')
        IGNOREORIGVCHDATE.text = 'No'
        ISSERVICETAXOVERRIDDEN = et.SubElement(VOUCHER, 'ISSERVICETAXOVERRIDDEN')
        ISSERVICETAXOVERRIDDEN.text = 'No'
        ISISDVOUCHER = et.SubElement(VOUCHER, 'ISISDVOUCHER')
        ISISDVOUCHER.text = 'No'
        ISEXCISEOVERRIDDEN = et.SubElement(VOUCHER, 'ISEXCISEOVERRIDDEN')
        ISEXCISEOVERRIDDEN.text = 'No'
        ISEXCISESUPPLYVCH = et.SubElement(VOUCHER, 'ISEXCISESUPPLYVCH')
        ISEXCISESUPPLYVCH.text = 'No'
        ISGSTOVERRIDDEN = et.SubElement(VOUCHER, 'ISGSTOVERRIDDEN')
        ISGSTOVERRIDDEN.text = 'No'
        GSTNOTEXPORTED = et.SubElement(VOUCHER, 'GSTNOTEXPORTED')
        GSTNOTEXPORTED.text = 'No'
        ISVATPRINCIPALACCOUNT = et.SubElement(VOUCHER, 'ISVATPRINCIPALACCOUNT')
        ISVATPRINCIPALACCOUNT.text = 'No'
        ISBOENOTAPPLICABLE = et.SubElement(VOUCHER, 'ISBOENOTAPPLICABLE')
        ISBOENOTAPPLICABLE.text = 'No'
        ISSHIPPINGWITHINSTATE = et.SubElement(VOUCHER, 'ISSHIPPINGWITHINSTATE')
        ISSHIPPINGWITHINSTATE.text = 'No'
        ISCANCELLED = et.SubElement(VOUCHER, 'ISCANCELLED')
        ISCANCELLED.text = 'No'
        HASCASHFLOW = et.SubElement(VOUCHER, 'HASCASHFLOW')
        HASCASHFLOW.text = 'No'
        ISPOSTDATED = et.SubElement(VOUCHER, 'ISPOSTDATED')
        ISPOSTDATED.text = 'No'
        USETRACKINGNUMBER = et.SubElement(VOUCHER, 'USETRACKINGNUMBER')
        USETRACKINGNUMBER.text = 'No'
        ISINVOICE = et.SubElement(VOUCHER, 'ISINVOICE')
        ISINVOICE.text = 'Yes'
        MFGJOURNAL = et.SubElement(VOUCHER, 'MFGJOURNAL')
        MFGJOURNAL.text = 'No'
        HASDISCOUNTS = et.SubElement(VOUCHER, 'HASDISCOUNTS')
        HASDISCOUNTS.text = 'Yes'
        ASPAYSLIP = et.SubElement(VOUCHER, 'ASPAYSLIP')
        ASPAYSLIP.text = 'No'
        ISCOSTCENTRE = et.SubElement(VOUCHER, 'ISCOSTCENTRE')
        ISCOSTCENTRE.text = 'Yes'
        ISSTXNONREALIZEDVCH = et.SubElement(VOUCHER, 'ISSTXNONREALIZEDVCH')
        ISSTXNONREALIZEDVCH.text = 'No'
        ISEXCISEMANUFACTURERON = et.SubElement(VOUCHER, 'ISEXCISEMANUFACTURERON')
        ISEXCISEMANUFACTURERON.text = 'No'
        ISBLANKCHEQUE = et.SubElement(VOUCHER, 'ISBLANKCHEQUE')
        ISBLANKCHEQUE.text = 'No'
        ISVOID = et.SubElement(VOUCHER, 'ISVOID')
        ISVOID.text = 'No'
        ISONHOLD = et.SubElement(VOUCHER, 'ISONHOLD')
        ISONHOLD.text = 'No'
        ORDERLINESTATUS = et.SubElement(VOUCHER, 'ORDERLINESTATUS')
        ORDERLINESTATUS.text = 'No'
        VATISAGNSTCANCSALES = et.SubElement(VOUCHER, 'VATISAGNSTCANCSALES')
        VATISAGNSTCANCSALES.text = 'No'
        VATISPURCEXEMPTED = et.SubElement(VOUCHER, 'VATISPURCEXEMPTED')
        VATISPURCEXEMPTED.text = 'No'
        ISVATRESTAXINVOICE = et.SubElement(VOUCHER, 'ISVATRESTAXINVOICE')
        ISVATRESTAXINVOICE.text = 'No'
        VATISASSESABLECALCVCH = et.SubElement(VOUCHER, 'VATISASSESABLECALCVCH')
        VATISASSESABLECALCVCH.text = 'No'
        ISVATDUTYPAID = et.SubElement(VOUCHER, 'ISVATDUTYPAID')
        ISVATDUTYPAID.text = 'Yes'
        ISDELIVERYSAMEASCONSIGNEE = et.SubElement(VOUCHER, 'ISDELIVERYSAMEASCONSIGNEE')
        ISDELIVERYSAMEASCONSIGNEE.text = 'No'
        ISDISPATCHSAMEASCONSIGNOR = et.SubElement(VOUCHER, 'ISDISPATCHSAMEASCONSIGNOR')
        ISDISPATCHSAMEASCONSIGNOR.text = 'No'
        ISDELETED = et.SubElement(VOUCHER, 'ISDELETED')
        ISDELETED.text = 'No'
        CHANGEVCHMODE = et.SubElement(VOUCHER, 'CHANGEVCHMODE')
        CHANGEVCHMODE.text = 'No'
        c = et.SubElement(VOUCHER, 'LEDGERENTRIES.LIST', TYPE='Ledger_Party')
        d = et.SubElement(c, 'OLDAUDITENTRYIDS.LIST', TYPE='Number')
        OLDAUDITENTRYIDS = et.SubElement(d, 'OLDAUDITENTRYIDS')
        OLDAUDITENTRYIDS.text = '-1'
        LEDGERNAME = et.SubElement(c, 'LEDGERNAME')
        LEDGERNAME.text = row[5]
        ISDEEMEDPOSITIVE = et.SubElement(c, 'ISDEEMEDPOSITIVE')
        ISDEEMEDPOSITIVE.text = 'No'
        LEDGERFROMITEM = et.SubElement(c, 'LEDGERFROMITEM')
        LEDGERFROMITEM.text = 'No'
        REMOVEZEROENTRIES = et.SubElement(c, 'REMOVEZEROENTRIES')
        REMOVEZEROENTRIES.text = 'No'
        ISPARTYLEDGER = et.SubElement(c, 'ISPARTYLEDGER')
        ISPARTYLEDGER.text = 'Yes'
        ISLASTDEEMEDPOSITIVE = et.SubElement(c, 'ISLASTDEEMEDPOSITIVE')
        ISLASTDEEMEDPOSITIVE.text = 'no'
        AMOUNT = et.SubElement(c, 'AMOUNT')
        s = 0
        for m in range(1,31):
            n = 'Debit Ledger ' + str(m)+' Amount'
            if acttest(n) == True:
                s = s + float(row[headers[n]])
        AMOUNT.text = str(round(float(s)))
        VATEXPAMOUNT = et.SubElement(c, 'VATEXPAMOUNT')
        VATEXPAMOUNT.text = str(round(float(s)))
        d = et.SubElement(c, 'BILLALLOCATIONS.LIST')
        NAME = et.SubElement(d, 'NAME')
        NAME.text = row[headers['Supplier Inv No']]
        if acttest('Debit Period') == True:
            BILLCREDITPERIOD = et.SubElement(d, 'BILLCREDITPERIOD', JD='43467', P=row[headers['Debit Period']])
            BILLCREDITPERIOD.text = row[headers['Debit Period']]
        BILLTYPE = et.SubElement(d, 'BILLTYPE')
        BILLTYPE.text = 'New Ref'
        TDSDEDUCTEEISSPECIALRATE = et.SubElement(d, 'TDSDEDUCTEEISSPECIALRATE')
        TDSDEDUCTEEISSPECIALRATE.text = 'No'
        AMOUNT = et.SubElement(d, 'AMOUNT')
        AMOUNT.text = str(s)
        for j in range(1, 31):
            n1 = 'Debit Ledger ' + str(j)
            n2 = 'Debit Ledger ' + str(j) + ' Amount'
            n3 = 'DebitLedger' + str(j)
            #print('Value of n1 and n2 is:', n1 , ' and ', n2)
            #print('key in hader and row', headers[n1], row[headers[n1]], row[headers[n2]])
            #print('acttest result', acttest(n1),' and ', acttest(n2))
            if acttest(n1) == True and acttest(n2) == True: 
                f = et.SubElement(VOUCHER, 'LEDGERENTRIES.LIST' , TYPE=n3)
                e = et.SubElement(f, 'OLDAUDITENTRYIDS.LIST', TYPE='Number')
                OLDAUDITENTRYIDS = et.SubElement(e, 'OLDAUDITENTRYIDS')
                OLDAUDITENTRYIDS.text = '-1'
                LEDGERNAME = et.SubElement(f, 'LEDGERNAME')
                LEDGERNAME.text = row[headers[n1]]
                if acttest('Tax Classification') == True and j < 2:
                    GSTOVRDNNATURE = et.SubElement(f, 'GSTOVRDNNATURE')
                    GSTOVRDNNATURE.text = row[headers['Tax Classification']]
                ISDEEMEDPOSITIVE = et.SubElement(f, 'ISDEEMEDPOSITIVE')
                ISDEEMEDPOSITIVE.text = 'Yes'
                LEDGERFROMITEM = et.SubElement(f, 'LEDGERFROMITEM')
                LEDGERFROMITEM.text = 'No'
                REMOVEZEROENTRIES = et.SubElement(f, 'REMOVEZEROENTRIES')
                REMOVEZEROENTRIES.text = 'No'
                ISPARTYLEDGER = et.SubElement(f, 'ISPARTYLEDGER')
                ISPARTYLEDGER.text = 'No'
                ISLASTDEEMEDPOSITIVE = et.SubElement(f, 'ISLASTDEEMEDPOSITIVE')
                ISLASTDEEMEDPOSITIVE.text = 'Yes'
                AMOUNT = et.SubElement(f, 'AMOUNT')
                AMOUNT.text = ""+str(-round(float(row[headers[n2]])))
                if acttest(n2) == 0:
                    VOUCHER.remove(f)
                    next
                VATEXPAMOUNT = et.SubElement(f, 'VATEXPAMOUNT')
                VATEXPAMOUNT.text = ""+str(-round(float(row[headers[n2]])))
                g = et.SubElement(f, 'RATEDETAILS.LIST')
                GSTRATEDUTYHEAD = et.SubElement(g, 'GSTRATEDUTYHEAD')
                GSTRATEDUTYHEAD.text = 'Integrated Tax'
                GSTRATEVALUATIONTYPE = et.SubElement(g, 'GSTRATEVALUATIONTYPE')
                GSTRATEVALUATIONTYPE.text = 'Based on Value'
                if acttest('GSTRate') == True and m < 2:                 
                    GSTRATE = et.SubElement(g, 'GSTRATE')
                    GSTRATE.text = str(row[headers['GSTRate']])
                g = et.SubElement(f, 'RATEDETAILS.LIST')
                GSTRATEDUTYHEAD = et.SubElement(g, 'GSTRATEDUTYHEAD')
                GSTRATEDUTYHEAD.text = 'Central Tax'
                GSTRATEVALUATIONTYPE = et.SubElement(g, 'GSTRATEVALUATIONTYPE')
                GSTRATEVALUATIONTYPE.text = 'Based on Value'
                if acttest('GSTRate') == True and j < 2:
                    GSTRATE = et.SubElement(g, 'GSTRATE')
                    GSTRATE.text = str(float(row[headers['GSTRate']])/2)
                g = et.SubElement(f, 'RATEDETAILS.LIST')
                GSTRATEDUTYHEAD = et.SubElement(g, 'GSTRATEDUTYHEAD')
                GSTRATEDUTYHEAD.text = 'State Tax'
                GSTRATEVALUATIONTYPE = et.SubElement(g, 'GSTRATEVALUATIONTYPE')
                GSTRATEVALUATIONTYPE.text = 'Based on Value'
                if acttest('GSTRate') == True and j < 2:
                    GSTRATE = et.SubElement(g, 'GSTRATE')
                    GSTRATE.text = str(float(row[headers['GSTRate']])/2)
                g = et.SubElement(f, 'RATEDETAILS.LIST')
                GSTRATEDUTYHEAD = et.SubElement(g, 'GSTRATEDUTYHEAD')
                GSTRATEDUTYHEAD.text = 'Cess'
                GSTRATEVALUATIONTYPE = et.SubElement(g, 'GSTRATEVALUATIONTYPE')
                GSTRATEVALUATIONTYPE.text = 'Based on Value'

tree = et.ElementTree(root)

def indent(elem, level=0):
    i = "\n" + level*"  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

indent(root)

with open ('with_breaks.xml', 'wb') as file:
    tree.write(file)
      
